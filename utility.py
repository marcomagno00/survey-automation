from dataclasses import dataclass, fields
from openpyxl import Workbook
import csv


def write_dataclass_to_excel(dataclass_obj, filename):
    
    # Create a new workbook and get the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Write headers (the alias if present, otherwise the field name)
    for col_index, field_info in enumerate(fields(dataclass_obj), start=1):
        alias = field_info.metadata.get("alias", field_info.name)
        worksheet.cell(row=1, column=col_index, value=alias)
        
    # Write the dataclass object's values
    for col_index, field_info in enumerate(fields(dataclass_obj), start=1):
        value = getattr(dataclass_obj, field_info.name)
        worksheet.cell(row=2, column=col_index, value=value)

    # Save the workbook
    workbook.save(filename)

def write_dataclass_to_csv(dataclass_obj, filename):
    # Extract aliases (or field names if no alias)
    header = []
    row = []
    for field_info in fields(dataclass_obj):
        alias = field_info.metadata.get("alias", field_info.name)
        header.append(alias)
        row.append(getattr(dataclass_obj, field_info.name))
    
    # Write to CSV
    with open(filename, mode="w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(header)  # Write header row
        writer.writerow(row)     # Write data row