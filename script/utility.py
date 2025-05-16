from dataclasses import dataclass, fields
from openpyxl import Workbook
import csv
import os
import shutil
import random

# Questo file Python contiene funzioni utilizzate nel main.py:
# funzioni per la gestione di file, immagini e cartelle
# funzioni per la scrittura di file Excel e TSV


def rename_images_in_folder(folder_path, base_name):
    image_extensions = ('.jpg', '.jpeg', '.png')
    
    try:
        files = sorted(os.listdir(folder_path))
        image_files = [f for f in files if f.lower().endswith(image_extensions)]

        for idx, filename in enumerate(image_files, start=1):
            ext = os.path.splitext(filename)[1]
            new_name = f"{base_name}{idx}{ext}"
            old_path = os.path.join(folder_path, filename)
            new_path = os.path.join(folder_path, new_name)
            os.rename(old_path, new_path)
            print(f"Renamed '{filename}' to '{new_name}'")
    except FileNotFoundError:
        print(f"Folder not found: {folder_path}")
    except Exception as e:
        print(f"Error: {e}")

def copy_images_to_folder(source_folder, destination_folder):
    image_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')

    try:
        os.makedirs(destination_folder, exist_ok=True)
        for root, _, files in os.walk(source_folder):
            for fname in files:
                if fname.lower().endswith(image_extensions):
                    dest_path = os.path.join(destination_folder, fname)
                    if os.path.exists(dest_path):
                        continue

                    src_path = os.path.join(root, fname)
                    shutil.copy2(src_path, dest_path)
    except FileNotFoundError:
        print(f"Folder not found: {source_folder}")
    except Exception as e:
        print(f"Error: {e}")

def count_images_in_folder(folder_path):
    # Common image extensions
    image_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')
    
    try:
        files = os.listdir(folder_path)
        image_files = [file for file in files if file.lower().endswith(image_extensions)]
        return len(image_files)
    except FileNotFoundError:
        print(f"Folder not found: {folder_path}")
        return 0

def get_images_from_folder(folder_path):
    # Estensioni tipiche delle immagini
    image_extensions = ('.jpg', '.jpeg', '.png')

    # Lista per contenere i nomi delle immagini
    image_names = []

    # Controlla se la cartella esiste
    if not os.path.isdir(folder_path):
        raise ValueError(f"La cartella '{folder_path}' non esiste.")

    # Itera attraverso i file nella cartella
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(image_extensions):
            image_names.append(filename)

    return image_names

def unify_and_randomize_images(cartella_corrette, cartella_sbagliate, output_tsv):
    # Recupera i path delle immagini
    immagini_corrette = [
        os.path.join(cartella_corrette, f) for f in os.listdir(cartella_corrette)
        if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))
    ]
    immagini_sbagliate = [
        os.path.join(cartella_sbagliate, f) for f in os.listdir(cartella_sbagliate)
        if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))
    ]

    # Associa etichetta 1 a corrette e 0 a sbagliate
    dati = [(path, 1) for path in immagini_corrette] + [(path, 0) for path in immagini_sbagliate]

    # Mescola i dati
    random.shuffle(dati)

    # Scrivi in un file TSV
    with open(output_tsv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerow(['path', 'label'])  # intestazione
        for path, label in dati:
            writer.writerow([path, label])

    print(f"File TSV creato con successo: {output_tsv}")

def give_first_txt_from_folder(folder_path):
    # Get the first file from the folder
    for file_name in os.listdir(folder_path):
        # check if file ends with .txt
        if file_name.endswith(".txt"):
            # return the first .txt file found
            return os.path.join(folder_path, file_name)
    return None

def read_first_line_from_txt(file_path):
    if file_path is None:
        return None
    # Open the file and read the first line
    with open(file_path, "r", encoding="utf-8") as file:
        first_line = file.readline().strip()
    return first_line

def read_all_lines_from_txt(file_path):
    lines = []
    if file_path is None:
        return None
    # Open the file and read the first line
    with open(file_path, "r", encoding="utf-8") as file:
        for line in file.readlines():
            line = line.strip()
            if line != "":
                lines.append(line)
    return lines

def write_dataclass_to_excel(dataclass_obj, filename):
    
    # Create a new workbook and get the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Write headers (the alias if present, otherwise the field name)
    for col_index, field_info in enumerate(fields(dataclass_obj), start=1):
        alias = field_info.metadata.get("alias", field_info.name)
        worksheet.cell(row=1, column=col_index, value=alias)
        
    # Write the dataclass object's values
    for col_index, field_info in enumerate(fields(dataclass_obj), start=1):
        value = getattr(dataclass_obj, field_info.name)
        worksheet.cell(row=2, column=col_index, value=value)

    # Save the workbook
    workbook.save(filename)

def write_dataclass_to_tsv(dataclass_obj, filename):
    # Extract aliases (or field names if no alias)
    header = []
    row = []
    for field_info in fields(dataclass_obj):
        alias = field_info.metadata.get("alias", field_info.name)
        header.append(alias)
        row.append(getattr(dataclass_obj, field_info.name))
    
    # Write to CSV
    with open(filename, mode="w", newline="", encoding="utf-8") as tsv_file:
        writer = csv.writer(tsv_file, delimiter="\t")
        writer.writerow(header)  # Write header row
        writer.writerow(row)     # Write data row

def add_row_to_tsv(dataclass_obj, filename):

    # Open the TSV file in append mode
    with open(filename, mode='a', newline='', encoding="utf-8") as file:
        writer = csv.writer(file, delimiter="\t")

        # If the file is empty, write the headers first
        if file.tell() == 0:
            headers = [field_info.metadata.get("alias", field_info.name) for field_info in fields(dataclass_obj)]
            writer.writerow(headers)

        # Write the values of the dataclass instance as a new row
        row = [getattr(dataclass_obj, field.name) for field in fields(dataclass_obj)]
        writer.writerow(row)

def convert_tsv_to_txt(input_path, output_path, delimiter="\t"):
    try:
        with open(input_path, 'r', encoding='utf-8', errors='replace') as infile, \
             open(output_path, 'w', encoding='utf-8') as outfile:

            for line in infile:
                # Rimuove newline finali e suddivide la riga su tab
                fields = line.rstrip('\n').split('\t')
                # Unisce i campi con il delimitatore scelto e scrive su file
                outfile.write(delimiter.join(fields) + '\n')

        print(f"Conversione completata: '{output_path}'")
    except Exception as e:
        print(f"Errore durante la conversione: {e}")

def create_survey(survey_settings, survey_cases, file_path_output):
    # settings
    for item in survey_settings:
        add_row_to_tsv(item, file_path_output)

    # cases 
    for case in survey_cases:
       for group_ in case:
            add_row_to_tsv(group_, file_path_output)

    print(f"File creato: '{file_path_output}'")